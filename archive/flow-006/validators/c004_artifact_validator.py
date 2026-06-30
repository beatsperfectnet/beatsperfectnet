#!/usr/bin/env python3
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
WORKBOOK = REPO / "archive/candidates/C-004-001/C-004-001-FAILED/product/Inventory-Tracker-Studio.xlsx"


def fail(message: str) -> None:
    raise AssertionError(message)


def non_formula_input_rows(ws, start: int, end: int, columns: tuple[int, ...]) -> int:
    count = 0
    for row in range(start, end + 1):
        has_value = False
        for col in columns:
            value = ws.cell(row, col).value
            if value not in (None, "") and not (isinstance(value, str) and value.startswith("=")):
                has_value = True
                break
        if has_value:
            count += 1
    return count


def formula_rows(ws, start: int, end: int) -> int:
    count = 0
    for row in range(start, end + 1):
        if any(isinstance(ws.cell(row, col).value, str) and ws.cell(row, col).value.startswith("=") for col in range(1, ws.max_column + 1)):
            count += 1
    return count


def sheet_text(ws) -> str:
    values: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                values.append(cell.value)
    return "\n".join(values).lower()


def validate() -> None:
    if not WORKBOOK.exists():
        fail(f"Workbook missing: {WORKBOOK}")

    wb = load_workbook(WORKBOOK, data_only=False)
    required = ["Start Here", "Dashboard", "Setup", "Inventory", "Purchases", "Sales", "Price Calculator", "Price List", "Help"]
    missing = [name for name in required if name not in wb.sheetnames]
    if missing:
        fail(f"Missing workbook sheets: {missing}")

    inventory = wb["Inventory"]
    headers = [inventory.cell(4, col).value for col in range(1, 22)]
    if "Reorder Point" in headers:
        fail("Inventory still exposes Reorder Point as a buyer-facing input header")
    if "Derived Reorder Point" not in headers:
        fail("Inventory must expose Derived Reorder Point as protected formula output")

    expected_seed_counts = {
        "Inventory": (8, wb["Inventory"], 5, 104, (1, 2, 3, 4, 5, 10, 21)),
        "Purchases": (16, wb["Purchases"], 5, 304, (1, 2, 3, 4, 5, 6, 8)),
        "Sales": (24, wb["Sales"], 5, 504, (1, 2, 3, 4, 5, 6, 10)),
        "Price Calculator": (8, wb["Price Calculator"], 5, 104, (1, 2, 4, 5, 6, 7, 8, 9, 13)),
        "Price List": (8, wb["Price List"], 5, 104, (1, 2, 3, 9)),
    }
    for name, (expected, ws, start, end, columns) in expected_seed_counts.items():
        actual = non_formula_input_rows(ws, start, end, columns)
        if actual != expected:
            fail(f"{name} must have {expected} seeded input rows, found {actual}")
        formulas = formula_rows(ws, start, end)
        if formulas != end - start + 1:
            fail(f"{name} must preserve formulas over all capacity rows, found {formulas}")

    if inventory["I5"].protection.locked is not True:
        fail("Derived reorder point column must be locked")
    for cell_ref in ("A13", "E13", "J13"):
        if inventory[cell_ref].protection.locked:
            fail(f"Blank working input cell {cell_ref} must be editable")
    if wb["Setup"]["B10"].protection.locked is not True:
        fail("Setup reorder point policy note must not be an editable buyer input")

    i_formula = str(inventory["I5"].value or "")
    if "ROUNDUP(L5*O5+P5,0)" not in i_formula:
        fail("Derived reorder point must be calculated from velocity, lead time, and safety stock")
    if "Fix stock data" not in str(inventory["S5"].value or ""):
        fail("Reorder action must block negative stock with Fix stock data")
    if "Review stock data" not in str(inventory["T5"].value or ""):
        fail("Status must flag negative stock as Review stock data")

    purchases = wb["Purchases"]
    sales = wb["Sales"]
    purchase_qty = defaultdict(float)
    sales_qty = defaultdict(float)
    for row in range(5, 305):
        sku = purchases.cell(row, 2).value
        qty = purchases.cell(row, 4).value or 0
        if sku:
            purchase_qty[sku] += qty
    for row in range(5, 505):
        sku = sales.cell(row, 2).value
        qty = sales.cell(row, 4).value or 0
        if sku:
            sales_qty[sku] += qty
    negatives = []
    for row in range(5, 105):
        sku = inventory.cell(row, 1).value
        if not sku:
            continue
        opening = inventory.cell(row, 5).value or 0
        raw_on_hand = opening + purchase_qty[sku] - sales_qty[sku]
        if raw_on_hand < 0:
            negatives.append((row, sku, raw_on_hand))
    if negatives:
        fail(f"Seeded artifact contains negative on-hand rows: {negatives}")

    start_text = sheet_text(wb["Start Here"])
    help_text = sheet_text(wb["Help"])
    if "reorder points" in start_text or "reorder points" in help_text:
        fail("Buyer-facing helper text must not tell buyers to fill reorder points")
    for phrase in ("blank", "fix stock data", "derived reorder point"):
        if phrase not in start_text and phrase not in help_text:
            fail(f"Start Here/Help must explain {phrase!r}")


if __name__ == "__main__":
    try:
        validate()
    except Exception as exc:
        print(f"C-004 artifact validation failed: {exc}", file=sys.stderr)
        sys.exit(1)
    print("C-004 artifact validation passed")
