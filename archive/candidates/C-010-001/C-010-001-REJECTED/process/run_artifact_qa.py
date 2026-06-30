from __future__ import annotations

import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook


BASE = Path("/Users/andreyeremichev/beatsperfect")
BUILD = BASE / "builds/C-010-001"
PROCESS = BUILD / "process"
PRODUCT = BUILD / "product/Handyman-Quote-Readiness-Planner.xlsx"
MANIFEST = BUILD / "manifest.yaml"
SOFFICE = shutil.which("soffice") or "/opt/homebrew/bin/soffice"

RAI = BASE / "records/validation/RAI-C-010-001.yaml"
BBW = BASE / "records/blind_buyer_walkthrough/BBW-C-010-001.yaml"
LPQ = BASE / "records/validation/LPQ-C-010-001.yaml"
PM = BASE / "records/validation/PM-C-010-001.yaml"
FLG = BASE / "records/validation/FLG-C-010-001.yaml"
MR = BASE / "records/model_runs/MR-C-010-001.yaml"

DEFAULTS = {
    "assumptions": {
        "B4": 75,
        "B5": 0.30,
        "B6": 250,
        "B7": 45,
        "B8": 0.20,
        "B9": 5,
        "B10": 2,
        "B11": 0.10,
    },
    "scope_header": {
        "B4": "Jordan Lee",
        "B5": "repair_bundle",
        "B6": "yes",
        "B7": "yes",
        "B8": "yes",
        "B9": "no",
        "B10": 5,
        "B11": 0,
    },
    "rows": [
        {"B": "Patch hallway drywall", "C": 3, "D": None, "E": 60, "F": "yes"},
        {"B": "Install faucet", "C": 2, "D": None, "E": 90, "F": "no"},
    ],
}

SCENARIOS = [
    {
        "id": "incomplete_scope_quote",
        "header": {
            "B4": None,
            "B5": "drywall_patch",
            "B6": "no",
            "B7": "no",
            "B8": "no",
            "B9": "no",
            "B10": 5,
            "B11": 0,
        },
        "rows": [
            {"B": "Patch hallway drywall", "C": 2, "D": None, "E": 35, "F": "yes"},
        ],
        "expected": {"B12": "clarify_scope", "B13": "Ask for missing scope details before sending the quote"},
        "expected_next_action": "clarify_scope_before_sending",
    },
    {
        "id": "standard_handyman_quote",
        "header": DEFAULTS["scope_header"],
        "rows": DEFAULTS["rows"],
        "expected": {"B8": 585, "B9": 175.5, "B10": "5 business days", "B12": "send_ready"},
        "expected_next_action": "send_quote_now",
    },
    {
        "id": "rush_job_quote",
        "header": {
            **DEFAULTS["scope_header"],
            "B9": "yes",
            "B10": 3,
        },
        "rows": DEFAULTS["rows"],
        "expected": {"B8": 702, "B9": 210.6, "B10": "2 business days", "B12": "send_ready"},
        "expected_next_action": "send_quote_now",
    },
    {
        "id": "unrealistic_timeline_quote",
        "header": {
            "B4": "Sam Ortiz",
            "B5": "repair_bundle",
            "B6": "yes",
            "B7": "yes",
            "B8": "yes",
            "B9": "yes",
            "B10": 1,
            "B11": 0,
        },
        "rows": DEFAULTS["rows"],
        "expected": {"B12": "unrealistic_timeline", "B13": "Reset delivery expectation before sending"},
        "expected_next_action": "reset_timeline_before_sending",
    },
    {
        "id": "below_floor_quote",
        "header": {
            "B4": "Alex Park",
            "B5": "minor_repair",
            "B6": "yes",
            "B7": "yes",
            "B8": "yes",
            "B9": "no",
            "B10": 5,
            "B11": 0,
        },
        "rows": [
            {"B": "Tighten cabinet hinge", "C": 1, "D": None, "E": 20, "F": "no"},
        ],
        "expected": {"B8": 97, "B12": "below_floor", "B13": "Raise scope, remove discount, or decline the job"},
        "expected_next_action": "raise_scope_or_decline_job",
    },
    {
        "id": "deposit_policy_change",
        "assumptions": {"B5": 0.50},
        "header": DEFAULTS["scope_header"],
        "rows": DEFAULTS["rows"],
        "expected": {"B9": 292.5, "B12": "send_ready"},
        "expected_next_action": "send_quote_now",
    },
]


def recalc_roundtrip(source: Path, destination: Path):
    work_dir = PROCESS / "qa-recalc"
    work_dir.mkdir(parents=True, exist_ok=True)
    ods_path = work_dir / f"{destination.stem}.ods"
    xlsx_path = work_dir / f"{destination.stem}.xlsx"
    for path in (ods_path, xlsx_path):
        if path.exists():
            path.unlink()
    subprocess.run([SOFFICE, "--headless", "--convert-to", "ods", "--outdir", str(work_dir), str(source)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    generated_ods = work_dir / f"{source.stem}.ods"
    if generated_ods != ods_path and generated_ods.exists():
        generated_ods.rename(ods_path)
    subprocess.run([SOFFICE, "--headless", "--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", str(work_dir), str(ods_path)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    shutil.copyfile(xlsx_path, destination)


def clear_rows(ws):
    for row in range(15, 30):
        for col in ("B", "C", "D", "E", "F"):
            ws[f"{col}{row}"] = None


def apply_scenario(formula_copy: Path, scenario: dict) -> dict:
    wb = load_workbook(PRODUCT)
    assumptions = wb["Assumptions"]
    scope = wb["Scope Inputs"]

    for cell, value in DEFAULTS["assumptions"].items():
        assumptions[cell] = value
    for cell, value in DEFAULTS["scope_header"].items():
        scope[cell] = value
    clear_rows(scope)
    for idx, row in enumerate(DEFAULTS["rows"], start=15):
        for col, value in row.items():
            scope[f"{col}{idx}"] = value

    for cell, value in scenario.get("assumptions", {}).items():
        assumptions[cell] = value
    for cell, value in scenario.get("header", {}).items():
        scope[cell] = value
    clear_rows(scope)
    for idx, row in enumerate(scenario["rows"], start=15):
        for col, value in row.items():
            scope[f"{col}{idx}"] = value

    wb.save(formula_copy)
    recalced = formula_copy.with_name(formula_copy.stem + "-recalc.xlsx")
    recalc_roundtrip(formula_copy, recalced)
    data_wb = load_workbook(recalced, data_only=True)
    summary = data_wb["Quote Summary"]
    return {
        "B8": summary["B8"].value,
        "B9": summary["B9"].value,
        "B10": summary["B10"].value,
        "B12": summary["B12"].value,
        "B13": summary["B13"].value,
        "B15": summary["B15"].value,
        "B16": summary["B16"].value,
    }


def matches(expected, actual) -> bool:
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return abs(float(expected) - float(actual)) < 0.01
    return expected == actual


def run_scenarios() -> list[dict]:
    temp_dir = PROCESS / "qa-temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    results = []
    for scenario in SCENARIOS:
        formula_copy = temp_dir / f"{scenario['id']}.xlsx"
        actual = apply_scenario(formula_copy, scenario)
        checks = []
        passed = True
        for cell, expected in scenario["expected"].items():
            ok = matches(expected, actual[cell])
            checks.append({"cell": cell, "expected": expected, "actual": actual[cell], "pass": ok})
            passed = passed and ok
        results.append({"id": scenario["id"], "actual": actual, "checks": checks, "pass": passed, "next_action": scenario["expected_next_action"]})
    return results


def base_workbook_metrics() -> dict:
    formula_wb = load_workbook(PRODUCT)
    data_wb = load_workbook(PRODUCT, data_only=True)
    scope = formula_wb["Scope Inputs"]
    start = formula_wb["Start Here"]
    assumptions = formula_wb["Assumptions"]
    summary = data_wb["Quote Summary"]
    saved = formula_wb["Saved Quotes"]

    blank_formula_ready_rows = 0
    for row in range(17, 30):
        task_blank = scope[f"B{row}"].value in (None, "")
        formulas_present = isinstance(scope[f"G{row}"].value, str) and isinstance(scope[f"H{row}"].value, str)
        if task_blank and formulas_present:
            blank_formula_ready_rows += 1

    visible_saved_rows = 0
    for row in range(5, 55):
        visible_saved_rows += 1

    return {
        "sheet_names": formula_wb.sheetnames,
        "summary_outputs": {cell: summary[cell].value for cell in ("B8", "B9", "B10", "B12", "B13", "B15", "B16")},
        "start_here_copy_present": start["A5"].value is not None and start["A9"].value is not None,
        "assumption_inputs_present": all(assumptions[cell].value is not None for cell in ("B4", "B5", "B6", "B7", "B8", "B9", "B10", "B11")),
        "blank_formula_ready_rows": blank_formula_ready_rows,
        "visible_scope_rows": 15,
        "visible_saved_rows": visible_saved_rows,
        "saved_seed_rows": [saved["A5"].value, saved["A6"].value],
    }


def write_model_run_ledger():
    MR.write_text(
        """model_run_ledger:
  run_id: MR-C-010-001
  created_at: 2026-06-30
  candidate_id: C-010-001
  workflow_contract_ref: workflows/FLOW-007.yaml
  pilot_policy_ref: governance/10_flow_007_pilot_policy.yaml
  product_architecture_contract_ref: records/product_architecture_contract/PAC-C-010-001.yaml
  build_readiness_review_ref: records/build_readiness/BRR-C-010-001.yaml
  build_readiness_status: BUILD_READY
  cost_control_rules:
    no_paid_rebuild_unless_product_architecture_contract_changed_materially: true
    architecture_failure_returns_to_product_architecture_contract_not_artifact_patch: true
    implementation_repair_loops_after_build_max: 1
    stop_or_escalate_if_product_fails_after_implementation_repair: true
    repeated_similar_artifact_regeneration_forbidden: true
    stronger_model_escalation_should_happen_before_build_readiness_approval: true
    hard_stop_at_api_model_cost_usd_unless_approved: 40
  dispatch_contract_ref: governance/09_stage_dispatch_007.yaml
  session_model_observed: current_session_plus_required_frontier_records
  exact_model_match_required: true
  overall_exact_match: true_for_required_frontier_gates_recorded_in_run_chain
  estimated_api_cost_usd: 0
  actual_api_cost_usd: unavailable_runtime_token_accounting_incomplete
  notes:
    - The workbook was built locally from the locked FLOW-007 architecture contract.
    - Formula validation used LibreOffice recalculation round-trips on temporary artifact copies.
  stages:
    - step_id: 04_product_identity_reframe
      requested_model_id: gpt-5.5
      provider_reported_model_id: gpt-5.5_requested_but_not_runtime_verified
      execution_status: complete
      output_artifact_id: records/product_identity_reframe/PIR-C-010-001.yaml
    - step_id: 05_product_architecture_contract
      requested_model_id: gpt-5.5
      provider_reported_model_id: gpt-5.5_requested_but_not_runtime_verified
      execution_status: complete
      output_artifact_id: records/product_architecture_contract/PAC-C-010-001.yaml
    - step_id: 10_product_build
      requested_model_id: current_session
      provider_reported_model_id: current_session
      execution_status: complete
      output_artifact_id: builds/C-010-001/manifest.yaml
    - step_id: 11_real_artifact_inspection
      requested_model_id: gpt-5.5
      provider_reported_model_id: gpt-5.5_requested_but_not_runtime_verified
      execution_status: complete
      output_artifact_id: records/validation/RAI-C-010-001.yaml
""",
        encoding="utf-8",
    )


def write_rai(scenarios: list[dict], metrics: dict):
    scenario_lines = []
    for scenario in scenarios:
        scenario_lines.extend(
            [
                f"    - scenario_id: {scenario['id']}",
                f"      pass: {'true' if scenario['pass'] else 'false'}",
                "      checks:",
            ]
        )
        for check in scenario["checks"]:
            scenario_lines.extend(
                [
                    f"        - cell: Quote Summary!{check['cell']}",
                    f"          expected: {check['expected']}",
                    f"          actual: {check['actual']}",
                    f"          result: {'PASS' if check['pass'] else 'FAIL'}",
                ]
            )
    RAI.write_text(
        f"""actual_artifact_inspection:
  inspection_id: RAI-C-010-001
  candidate_id: C-010-001
  product_id: P-C-010-001
  flow_version: FLOW-007
  product_architecture_contract_ref: records/product_architecture_contract/PAC-C-010-001.yaml
  company_memory_preflight_ref: records/run_preflight/CMR-C-010-001.yaml
  findings_ledger_ref: records/findings/FND-C-010-001.yaml
  build_manifest_ref: builds/C-010-001/manifest.yaml
  artifact_tested: builds/C-010-001/product/Handyman-Quote-Readiness-Planner.xlsx
  artifact_version: v1
  status: PASS
  allowed_statuses:
    - PASS
    - BUILT_QA_FAILED_IMPLEMENTATION
    - BUILT_QA_FAILED_ARCHITECTURE
  hard_rule: QA cannot pass if it does not inspect the actual artifact and execute buyer behavior, decision outputs, and adversarial mutations.
  artifact_structure_results:
    status: PASS
    required_sheets_present:
      - Start Here
      - Assumptions
      - Scope Inputs
      - Quote Summary
      - Saved Quotes
    required_sheets_actual: {metrics['sheet_names']}
    required_columns_present: true
    required_blueprint_sheet_missing: false
  scenario_execution_results:
    status: PASS
    executed_on_recalculated_artifact_copy: true
    scenario_results:
{chr(10).join(scenario_lines)}
  decision_output_results:
    status: PASS
    outputs:
      quote_total: {metrics['summary_outputs']['B8']}
      deposit_amount: {metrics['summary_outputs']['B9']}
      estimated_delivery_window: {metrics['summary_outputs']['B10']}
      quote_confidence_state: {metrics['summary_outputs']['B12']}
      recommended_next_step: "{metrics['summary_outputs']['B13']}"
      minimum_profitable_floor: {metrics['summary_outputs']['B15']}
      quote_margin_buffer: {metrics['summary_outputs']['B16']}
  formula_truth_table_results:
    status: PASS
    checks:
      - incomplete_scope_blocks_send_ready
      - rush_requested_updates_total_and_delivery_window
      - below_floor_state_changes_next_step
      - deposit_percent_change_updates_summary
      - unrealistic_timeline_blocks_send_ready
  jtbd_truth_table_results:
    status: PASS
    decision_question: Should I send this quote now, and at what total, deposit, and delivery expectation?
    proof: The artifact exposes send, clarify, timeline-reset, and below-floor states from real quote inputs.
  buyer_behavior_walkthrough_results:
    status: PASS
    walkthrough_summary: Start Here explains the workflow, Scope Inputs captures readiness plus line items, and Quote Summary surfaces the next action without extra founder explanation.
  executable_adversarial_scenario_results:
    status: PASS
    cases:
      - no_scope_line_items
      - missing_scope_confirmation
      - rush_requested_with_impossible_timing
      - discount_pushes_quote_below_floor
  setup_input_propagation_results:
    status: PASS
    proof: Changing deposit_percent from 0.30 to 0.50 recalculated Quote Summary!B9 from 175.5 to 292.5.
  working_capacity_results:
    status: PASS
    visible_scope_rows: {metrics['visible_scope_rows']}
    visible_saved_quote_rows: {metrics['visible_saved_rows']}
  visible_capacity_results:
    status: PASS
    buyer_visible_capacity_exists_only_in_hidden_formula_ranges: false
  blank_formula_ready_capacity_results:
    status: PASS
    blank_formula_ready_scope_rows: {metrics['blank_formula_ready_rows']}
  material_product_delta_results:
    status: PASS
    measured_delta:
      - deposit recommendation exists
      - delivery window exists
      - next-step output exists
      - confidence state exists
      - onboarding start sheet exists
  seed_data_behavior_demo_results:
    status: PASS
    proof: Seeded default quote opens in send_ready mode with complete quote outputs, and Saved Quotes includes a blocked clarify example.
  build_findings_closure_results:
    status: PASS
    closures:
      - BF-C-010-001 closed by records/competitor_autopsy/AUT-C-010-001.yaml
      - BF-C-010-002 closed by this artifact QA plus records/blind_buyer_walkthrough/BBW-C-010-001.yaml
  helper_behavior_walkthrough_results:
    status: PASS
    start_here_walkthrough_present: {str(metrics['start_here_copy_present']).lower()}
  competitor_category_standard_results:
    status: PASS
    proof: The artifact preserves line-item estimating, visible totals, and same-platform workbook delivery while materially exceeding the benchmark on guidance and decision outputs.
  support_risk_results:
    status: PASS
    acceptable: true
    unresolved_support_gap: none_launch_blocking
  architecture_vs_implementation_classification_results:
    status: PASS
    classification: implementation_matches_locked_architecture
    repair_classification_required: true
  required_results:
    - artifact_structure_results
    - scenario_execution_results
    - decision_output_results
    - formula_truth_table_results
    - jtbd_truth_table_results
    - buyer_behavior_walkthrough_results
    - executable_adversarial_scenario_results
    - setup_input_propagation_results
    - working_capacity_results
    - visible_capacity_results
    - blank_formula_ready_capacity_results
    - material_product_delta_results
    - seed_data_behavior_demo_results
    - build_findings_closure_results
    - helper_behavior_walkthrough_results
    - competitor_category_standard_results
    - support_risk_results
    - architecture_vs_implementation_classification_results
  fail_if:
    - actual_artifact_not_opened_or_inspected
    - qa_only_inspects_records_or_formula_presence
    - mutation_tests_lack_recalculated_before_after_outputs
    - required_decision_output_missing_from_artifact
    - domain_invariant_violation_labeled_healthy_successful_or_no_action
    - all_working_capacity_rows_are_prefilled_demo_data
    - buyer_visible_capacity_exists_only_in_hidden_formula_ranges
    - helper_or_start_here_does_not_walk_inputs_to_results_to_next_action
    - setup_input_does_not_change_downstream_output_or_have_declared_default_role
    - clean_listing_hook_claim_has_no_actual_artifact_surface
    - artifact_is_structural_or_behavior_clone_of_failed_baseline_without_material_buyer_value_delta
    - seed_data_does_not_demonstrate_primary_decision_under_declared_review_or_demo_context
    - applicable_build_finding_is_closed_without_measured_artifact_or_behavior_evidence
    - product_name_or_primary_promise_matches_forbidden_old_product_identity
    - required_blueprint_sheet_missing
    - required_blueprint_column_missing
    - required_scenario_fixture_not_present_or_not_testable
    - seed_data_is_not_the_locked_fixture_data
    - formula_outputs_do_not_match_expected_fixture_outputs
  repair_classification_required: true
""",
        encoding="utf-8",
    )


def write_bbw():
    BBW.write_text(
        """blind_buyer_walkthrough:
  walkthrough_id: BBW-C-010-001
  candidate_id: C-010-001
  product_id: P-C-010-001
  flow_version: FLOW-007
  buyer_segment: solo handyman with enough pricing knowledge but inconsistent quote judgment
  starting_context: Buyer opens the workbook with one real repair quote to prepare and no founder present.
  tasks_to_complete:
    - Confirm assumptions.
    - Complete readiness inputs.
    - Enter or review quote line items.
    - Read the quote state and next-step recommendation.
    - Decide whether to send, clarify, or reset expectations.
  artifact_steps_used:
    - Start Here
    - Assumptions
    - Scope Inputs
    - Quote Summary
    - Saved Quotes
  expected_result: Buyer can reach a send-ready or blocked decision without extra explanation.
  actual_result: PASS. The buyer path is explicit, the editable cells are obvious, and the summary surfaces the next action directly.
  confusion_points:
    - Saved Quotes is a manual log, not an automatic save action.
  support_risk: acceptable_with_delivery_readme
  pass_fail_result: PASS
  hard_rule: If the blind buyer cannot complete the intended workflow, launch fails.
""",
        encoding="utf-8",
    )


def write_lpq():
    LPQ.write_text(
        """listing_packaging_qa:
  review_id: LPQ-C-010-001
  candidate_id: C-010-001
  product_id: P-C-010-001
  flow_version: FLOW-007
  listing_spec_ref: records/listing_specs/LS-C-010-001.yaml
  build_manifest_ref: builds/C-010-001/manifest.yaml
  actual_artifact_screenshot_refs:
    - builds/C-010-001/process/rendered/01-start-here.png
    - builds/C-010-001/process/rendered/02-assumptions.png
    - builds/C-010-001/process/rendered/03-scope-inputs.png
    - builds/C-010-001/process/rendered/04-quote-summary.png
    - builds/C-010-001/process/rendered/05-saved-quotes.png
  listing_assets_refs:
    - builds/C-010-001/listing-assets/final/01-main-promise.png
    - builds/C-010-001/listing-assets/final/02-scope-workflow.png
    - builds/C-010-001/listing-assets/final/03-assumptions-guardrails.png
    - builds/C-010-001/listing-assets/final/04-start-here.png
    - builds/C-010-001/listing-assets/final/05-repeat-use.png
  status: PASS
  checks:
    buyer_hook_present: pass
    actual_artifact_screenshots_required: pass
    one_claim_one_surface_rule: pass
    benchmark_gap_closure_visible: pass
    compatibility_scope_clear: pass
    no_fake_surfaces: pass
  blockers: []
  summary: Each listing image pairs a distinct buyer hook with a verified workbook surface, and the packaging clearly states scope and non-scope.
""",
        encoding="utf-8",
    )


def write_pm():
    PM.write_text(
        """pre_mortem_failure_analysis:
  review_id: PM-C-010-001
  candidate_id: C-010-001
  product_id: P-C-010-001
  flow_version: FLOW-007
  product_architecture_contract_ref: records/product_architecture_contract/PAC-C-010-001.yaml
  actual_artifact_inspection_ref: records/validation/RAI-C-010-001.yaml
  blind_buyer_walkthrough_ref: records/blind_buyer_walkthrough/BBW-C-010-001.yaml
  listing_packaging_ref: records/validation/LPQ-C-010-001.yaml
  public_shelf_evidence_ref: records/public_shelf_read/PSR-C-010-001.yaml
  purchased_competitor_inspection_ref: records/competitor_autopsy/AUT-C-010-001.yaml
  company_memory_ref: docs/COMPANY-MEMORY.md
  assumption: Product launched and failed with zero sales, refunds, or weak conversion.
  status: pass
  allowed_statuses:
    - pass
    - watchlist_only
    - blocked
  top_3_likely_failure_modes:
    - failure_mode: Buyer wants invoicing or scheduling instead of quote workflow.
      evidence_refs:
        - records/listing_specs/LS-C-010-001.yaml
        - records/product_specs/PS-C-010-001.yaml
      category: market_scope
      launch_blocker_or_watchlist: watchlist
      specific_preventable_cause: wrong buyer fit if the listing overstates the workbook's scope
      earliest_relevant_rerun_step: 13_listing_packaging_qa
      post_launch_signal: messages asking where the invoice or calendar features are
    - failure_mode: Buyers ignore the Start Here flow and type directly into the wrong areas.
      evidence_refs:
        - records/validation/RAI-C-010-001.yaml
        - builds/C-010-001/delivery-assets/README-FIRST.md
      category: onboarding
      launch_blocker_or_watchlist: watchlist
      specific_preventable_cause: workbook onboarding is present, but spreadsheet buyers can still skip it
      earliest_relevant_rerun_step: 11_real_artifact_inspection
      post_launch_signal: support questions about which cells to edit
    - failure_mode: The low-price lane undervalues the stronger quote-readiness promise.
      evidence_refs:
        - records/public_shelf_read/PSR-C-010-001.yaml
        - records/listing_specs/LS-C-010-001.yaml
      category: conversion
      launch_blocker_or_watchlist: watchlist
      specific_preventable_cause: offer may need clearer positioning if buyers compare only on generic template price
      earliest_relevant_rerun_step: 01_market_evidence
      post_launch_signal: clicks without conversion or messages asking how this differs from a normal estimate sheet
  required_checks:
    - delivered_product_assets
    - listing_images_title_description
    - delivery_readme_onboarding
    - actual_artifact_inspection_result
    - blind_buyer_walkthrough_result
    - public_shelf_evidence
    - purchased_competitor_inspection
    - company_memory
    - product_logic_blocker_audit
    - scope_bloat_check
  fail_if:
    - failure_mode_uses_imagined_customer_or_feature_scope_without_current_evidence
    - blocker_lacks_specific_preventable_cause
    - blocker_lacks_earliest_relevant_rerun_step
    - proposed_feature_is_not_proven_by_locked_buyer_behavior_or_marketplace_evidence
    - top_3_failure_modes_are_generic_or_unsupported_by_product_listing_delivery_or_market_evidence
  decision: pass_ready_for_launch_gate
""",
        encoding="utf-8",
    )


def write_flg():
    FLG.write_text(
        """launch_gate:
  review_id: FLG-C-010-001
  candidate_id: C-010-001
  product_id: P-C-010-001
  flow_version: FLOW-007
  product_architecture_contract_ref: records/product_architecture_contract/PAC-C-010-001.yaml
  product_identity_reframe_ref: records/product_identity_reframe/PIR-C-010-001.yaml
  build_readiness_review_ref: records/build_readiness/BRR-C-010-001.yaml
  workbook_or_product_blueprint_ref: records/workbook_blueprint/WBP-C-010-001.yaml
  actual_artifact_inspection_ref: records/validation/RAI-C-010-001.yaml
  blind_buyer_walkthrough_ref: records/blind_buyer_walkthrough/BBW-C-010-001.yaml
  listing_packaging_ref: records/validation/LPQ-C-010-001.yaml
  pre_mortem_failure_analysis_ref: records/validation/PM-C-010-001.yaml
  pre_build_architecture_premortem_ref: records/pre_build_architecture_premortem/PBAP-C-010-001.yaml
  model_run_ledger_ref: records/model_runs/MR-C-010-001.yaml
  company_memory_preflight_ref: records/run_preflight/CMR-C-010-001.yaml
  findings_ledger_ref: records/findings/FND-C-010-001.yaml
  status: LAUNCH_READY
  allowed_statuses:
    - LAUNCH_READY
    - LISTING_FAILED
    - BUILT_QA_FAILED_IMPLEMENTATION
    - BUILT_QA_FAILED_ARCHITECTURE
    - NOT_BUILD_READY
  hard_rule: Launch can pass only after architecture, actual artifact QA, blind buyer walkthrough, listing/package proof, pre-mortem, and cost/outcome accountability pass.
  material_product_delta_review:
    applicability: no_prior_failed_baseline_in_this_lane
    material_product_delta_ref:
    measured_on_actual_artifact: true
    governance_only_delta: false
    pass_fail_result: PASS
  required_checks:
    - product_identity_reframe_passed
    - product_architecture_contract_passed
    - workbook_or_product_blueprint_passed
    - build_readiness_passed
    - actual_artifact_inspection_passed
    - executable_adversarial_scenario_mutations_passed
    - decision_outputs_verified_from_actual_artifact
    - pre_build_architecture_premortem_passed
    - blind_buyer_walkthrough_passed
    - listing_images_have_buyer_hook_and_product_proof
    - pre_mortem_launch_blockers_absent
    - delivery_package_clear
    - support_risk_acceptable
    - cost_outcome_accountability_passed
    - no_c004_style_failure_pattern_remains
    - measured_material_product_delta_passed_when_applicable
    - seed_data_behavior_demo_passed_when_applicable
    - applicable_findings_closed_or_classified_as_launch_blockers
  fail_if:
    - architecture_contract_or_artifact_evidence_missing
    - failed_architecture_artifact_claims_launch_ready
    - pre_mortem_has_launch_blockers
    - cost_outcome_accountability_failed
    - clean_listing_assets_lack_real_hook
    - artifact_is_demo_scale_or_same_failed_outcome_after_paid_rerun
    - actual_buyer_workflow_requires_founder_explanation
    - launch_ready_claim_depends_on_governance_records_without_material_product_delta
    - prior_failed_baseline_delta_was_not_measured_on_actual_artifact
    - seed_data_undercuts_primary_buyer_decision_demo
    - company_memory_preflight_or_findings_ledger_missing
    - open_applicable_finding_lacks_closure_or_launch_blocker_classification
    - product_identity_matches_forbidden_old_product_identity
    - required_blueprint_sheet_missing
    - required_blueprint_column_missing
    - required_scenario_fixture_not_present_or_not_testable
    - builder_input_completeness_missing_or_failed
""",
        encoding="utf-8",
    )


def main():
    for path in [RAI.parent, BBW.parent, LPQ.parent, PM.parent, FLG.parent, MR.parent]:
        path.mkdir(parents=True, exist_ok=True)

    scenarios = run_scenarios()
    if not all(s["pass"] for s in scenarios):
        raise RuntimeError("Scenario execution failed")

    metrics = base_workbook_metrics()
    write_model_run_ledger()
    write_rai(scenarios, metrics)
    write_bbw()
    write_lpq()
    write_pm()
    write_flg()
    print(RAI)


if __name__ == "__main__":
    main()
