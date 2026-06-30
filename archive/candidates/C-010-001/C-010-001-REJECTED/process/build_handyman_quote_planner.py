from __future__ import annotations

import shutil
import subprocess
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


BASE = Path("/Users/andreyeremichev/beatsperfect")
BUILD = BASE / "builds/C-010-001"
PROCESS = BUILD / "process"
PRODUCT_DIR = BUILD / "product"
DELIVERY_DIR = BUILD / "delivery-assets"
SURFACE_XLSX_DIR = PROCESS / "surface-export-workbooks"
SURFACE_PDF_DIR = PROCESS / "surface-export-pdf"
SURFACE_THUMB_DIR = PROCESS / "surface-export-thumbnails"
RENDER_DIR = PROCESS / "rendered"

PRODUCT = PRODUCT_DIR / "Handyman-Quote-Readiness-Planner.xlsx"
IMPORT_COPY = PRODUCT_DIR / "Handyman-Quote-Readiness-Planner-google-sheets-import.xlsx"
FORMULA_SOURCE = PROCESS / "formula-source" / "Handyman-Quote-Readiness-Planner.xlsx"

SOFFICE = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
PROTECTION_PASSWORD = "QuoteReady"

BG = "F5F1EA"
SURFACE = "FFFDFC"
TEXT = "1E293B"
MUTED = "52606D"
PIPELINE = "8C3B28"
SUCCESS = "5C7A52"
WARNING = "D97706"
DANGER = "B42318"
INPUT = "FFF2CC"
OUTPUT = "DDEBF7"
CARD = "F8F4EE"
BORDER = "D7C7B4"
NEUTRAL = "EEF2F6"

THIN = Side(style="thin", color=BORDER)
MED = Side(style="medium", color=PIPELINE)

SURFACES = [
    ("Start Here", "01-start-here"),
    ("Assumptions", "02-assumptions"),
    ("Scope Inputs", "03-scope-inputs"),
    ("Quote Summary", "04-quote-summary"),
    ("Saved Quotes", "05-saved-quotes"),
]


def fill(rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=rgb)


def border_for(left=False, right=False, top=False, bottom=False) -> Border:
    return Border(
        left=MED if left else THIN,
        right=MED if right else THIN,
        top=MED if top else THIN,
        bottom=MED if bottom else THIN,
    )


def apply_grid(ws, start_row, end_row, start_col, end_col, *, cell_fill=None, font=None, align=None):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.border = border_for(
                left=col == start_col,
                right=col == end_col,
                top=row == start_row,
                bottom=row == end_row,
            )
            if cell_fill:
                cell.fill = fill(cell_fill)
            if font:
                cell.font = copy(font)
            if align:
                cell.alignment = copy(align)


def style_title(ws, title: str, subtitle: str, width: int = 8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = fill(PIPELINE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=10, italic=True, color=MUTED)
    ws["A2"].fill = fill(BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22


def label_value(ws, row: int, label_col: int, value_col: int, label: str, value, *, input_cell=False, output_cell=False):
    left = ws.cell(row, label_col, label)
    left.font = Font(name="Aptos", size=9, bold=True, color=MUTED)
    left.fill = fill(CARD)
    left.alignment = Alignment(horizontal="left", vertical="center")
    left.border = border_for(left=True, top=True, bottom=True)

    right = ws.cell(row, value_col, value)
    right.font = Font(name="Aptos", size=10, bold=input_cell or output_cell, color=TEXT)
    right.fill = fill(INPUT if input_cell else OUTPUT if output_cell else SURFACE)
    right.alignment = Alignment(horizontal="center" if output_cell else "left", vertical="center")
    right.border = border_for(right=True, top=True, bottom=True)
    ws.row_dimensions[row].height = 22


def unlock_range(ws, cell_range: str):
    for row in ws[cell_range]:
        for cell in row:
            cell.protection = Protection(locked=False)


def protect_sheet(ws):
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False
    ws.protection.sort = False


def set_dimensions(ws, widths: dict[str, float]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def yes_no_validation(ws, cell_range: str):
    dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=False)
    dv.errorTitle = "Use yes or no"
    dv.error = "Pick yes or no from the dropdown."
    dv.promptTitle = "Allowed values"
    dv.prompt = "Use yes or no."
    ws.add_data_validation(dv)
    dv.add(cell_range)


def decimal_validation(ws, cell_range: str):
    dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv.errorTitle = "Numbers only"
    dv.error = "Enter a number that is zero or higher."
    ws.add_data_validation(dv)
    dv.add(cell_range)


def build_start_here(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 90
    style_title(ws, "Handyman Quote Readiness Planner", "Quote the job, check the floor, and know whether to send or clarify before you reply.", width=7)
    set_dimensions(ws, {"A": 18, "B": 44, "C": 20, "D": 20, "E": 18, "F": 18, "G": 18})

    blocks = [
        ("Step 1", "Open Assumptions and confirm your hourly rate, minimum job floor, deposit policy, and standard timing."),
        ("Step 2", "Go to Scope Inputs and complete the readiness checks before touching the line items."),
        ("Step 3", "Enter task rows with hours, material cost, and travel yes/no. Yellow cells are yours."),
        ("Step 4", "Read Quote Summary. Send only when the state says send_ready. Otherwise follow the next-step guidance."),
    ]
    start_row = 5
    for index, (title, text) in enumerate(blocks):
        row = start_row + (index * 4)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
        ws.cell(row, 1).fill = fill(PIPELINE if index < 2 else SUCCESS if index == 2 else WARNING)
        ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 2, end_column=5)
        ws.cell(row + 1, 1, text)
        ws.cell(row + 1, 1).font = Font(name="Aptos", size=10, color=TEXT)
        ws.cell(row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
        apply_grid(ws, row, row + 2, 1, 5, cell_fill=SURFACE)

    ws["A22"] = "Color key"
    ws["A22"].font = Font(name="Aptos", size=11, bold=True, color=TEXT)
    color_key = [
        ("Yellow", "Buyer input cells"),
        ("Blue", "Generated outputs"),
        ("White", "Reference or helper text"),
    ]
    for idx, (label, desc) in enumerate(color_key, start=23):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, desc)
        ws.cell(idx, 1).font = Font(name="Aptos", size=10, bold=True, color=TEXT)
        ws.cell(idx, 2).font = Font(name="Aptos", size=10, color=TEXT)
        ws.cell(idx, 1).fill = fill(INPUT if label == "Yellow" else OUTPUT if label == "Blue" else SURFACE)
        apply_grid(ws, idx, idx, 1, 2)

    ws["D22"] = "This workbook is for quoting only."
    ws["D22"].font = Font(name="Aptos", size=11, bold=True, color=TEXT)
    notes = [
        "No invoicing, CRM, dispatch, or auto-emailing.",
        "If the state is clarify_scope, ask for more details before sending.",
        "If the state is below_floor, adjust scope or decline the job.",
        "If the state is unrealistic_timeline, reset expectations before you quote.",
    ]
    for idx, note in enumerate(notes, start=23):
        ws.cell(idx, 4, "\u2022 " + note)
        ws.cell(idx, 4).font = Font(name="Aptos", size=10, color=TEXT)
        ws.merge_cells(start_row=idx, start_column=4, end_row=idx, end_column=7)


def build_assumptions(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 90
    style_title(ws, "Assumptions", "Set the business defaults that every quote uses unless you override them.", width=5)
    set_dimensions(ws, {"A": 28, "B": 18, "C": 18, "D": 20, "E": 24})

    rows = [
        ("Default hourly rate", 75, "$/hour"),
        ("Deposit percent", 0.30, "percent"),
        ("Minimum job floor", 250, "$ minimum"),
        ("Travel fee per quoted job", 45, "$ flat"),
        ("Rush surcharge percent", 0.20, "percent"),
        ("Standard turnaround days", 5, "business days"),
        ("Rush turnaround days", 2, "business days"),
        ("Material markup percent", 0.10, "percent"),
    ]
    for offset, (label, value, meaning) in enumerate(rows, start=4):
        ws[f"A{offset}"] = label
        ws[f"B{offset}"] = value
        ws[f"C{offset}"] = meaning
        ws[f"A{offset}"].font = Font(name="Aptos", size=10, color=TEXT)
        ws[f"B{offset}"].font = Font(name="Aptos", size=10, bold=True, color=TEXT)
        ws[f"C{offset}"].font = Font(name="Aptos", size=10, color=MUTED)
        ws[f"B{offset}"].fill = fill(INPUT)
        ws[f"C{offset}"].fill = fill(SURFACE)
        ws[f"A{offset}"].fill = fill(SURFACE)
        apply_grid(ws, offset, offset, 1, 3)

    unlock_range(ws, "B4:B11")
    decimal_validation(ws, "B4:B11")

    ws["E4"] = "Guardrails"
    ws["E4"].font = Font(name="Aptos", size=11, bold=True, color=TEXT)
    helper_lines = [
        "Quote Summary uses these values every time.",
        "Change deposit percent here to update the summary.",
        "Minimum job floor blocks send_ready when totals fall below it.",
        "Rush timing and surcharge work together.",
    ]
    for idx, note in enumerate(helper_lines, start=5):
        ws[f"E{idx}"] = note
        ws[f"E{idx}"].font = Font(name="Aptos", size=10, color=TEXT)

    for row in range(4, 12):
        if row in (4, 6, 7):
            ws[f"B{row}"].number_format = '$#,##0.00'
        elif row in (5, 8, 11):
            ws[f"B{row}"].number_format = '0%'
        else:
            ws[f"B{row}"].number_format = '0'


def build_scope_inputs(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A14"
    ws.sheet_view.zoomScale = 85
    style_title(ws, "Scope Inputs", "Complete readiness checks first, then enter line items. Only yellow cells should be edited.", width=8)
    set_dimensions(ws, {"A": 6, "B": 28, "C": 14, "D": 14, "E": 14, "F": 14, "G": 16, "H": 16})

    header_rows = [
        ("Client name", "Jordan Lee"),
        ("Job type", "repair_bundle"),
        ("Site visit complete", "yes"),
        ("Material list complete", "yes"),
        ("Scope confirmed", "yes"),
        ("Rush requested", "no"),
        ("Requested days until start", 5),
        ("Discount amount", 0),
    ]
    row_num = 4
    for label, value in header_rows:
        label_value(ws, row_num, 1, 2, label, value, input_cell=True)
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=1)
        for col in range(1, 4):
            ws.cell(row_num, col).border = border_for(
                left=col == 1,
                right=col == 3,
                top=True,
                bottom=True,
            )
        row_num += 1

    yes_no_validation(ws, "B6:B9")
    decimal_validation(ws, "B10:B11")
    unlock_range(ws, "B4:B11")

    ws["A14"] = "No."
    ws["B14"] = "Task name"
    ws["C14"] = "Est. hours"
    ws["D14"] = "Rate override"
    ws["E14"] = "Material cost"
    ws["F14"] = "Travel required"
    ws["G14"] = "Labor subtotal"
    ws["H14"] = "Line subtotal"
    for col in range(1, 9):
        cell = ws.cell(14, col)
        cell.font = Font(name="Aptos", size=10, bold=True, color=TEXT)
        cell.fill = fill(CARD)
        cell.alignment = Alignment(horizontal="center")
    apply_grid(ws, 14, 14, 1, 8)

    seed_rows = [
        ("Patch hallway drywall", 3, None, 60, "yes"),
        ("Install faucet", 2, None, 90, "no"),
    ]
    for idx, row in enumerate(range(15, 30), start=1):
        ws[f"A{row}"] = idx
        if idx <= len(seed_rows):
            task, hours, rate_override, material_cost, travel = seed_rows[idx - 1]
            ws[f"B{row}"] = task
            ws[f"C{row}"] = hours
            ws[f"D{row}"] = rate_override
            ws[f"E{row}"] = material_cost
            ws[f"F{row}"] = travel
        ws[f"G{row}"] = f'=IF(B{row}="","",ROUND(C{row}*IF(D{row}="",Assumptions!$B$4,D{row}),2))'
        ws[f"H{row}"] = f'=IF(B{row}="","",ROUND(G{row}+(E{row}*(1+Assumptions!$B$11))+IF(F{row}="yes",Assumptions!$B$7,0),2))'
        for col in range(1, 9):
            cell = ws.cell(row, col)
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
            if col in (2, 3, 4, 5, 6):
                cell.fill = fill(INPUT)
            elif col in (7, 8):
                cell.fill = fill(OUTPUT)
            else:
                cell.fill = fill(SURFACE)
        apply_grid(ws, row, row, 1, 8)

    yes_no_validation(ws, "F15:F29")
    decimal_validation(ws, "C15:E29")
    unlock_range(ws, "B15:F29")

    for row in range(15, 30):
        for col in ("C", "D", "E", "G", "H"):
            ws[f"{col}{row}"].number_format = '$#,##0.00' if col in ("D", "E", "G", "H") else '0.00'
    ws["C15"].number_format = '0.00'
    ws["C16"].number_format = '0.00'
    for row in range(17, 30):
        ws[f"C{row}"].number_format = '0.00'

    ws["J4"] = "Helper totals"
    ws["J4"].font = Font(name="Aptos", size=11, bold=True, color=TEXT)
    helper = [
        ("Base total before rush", '=ROUND(SUM(H15:H29)-B11,2)'),
        ("Valid task rows", '=COUNTIF(B15:B29,"<>")'),
        ("Travel lines", '=COUNTIF(F15:F29,"yes")'),
    ]
    for idx, (label, formula) in enumerate(helper, start=5):
        ws[f"J{idx}"] = label
        ws[f"K{idx}"] = formula
        ws[f"J{idx}"].font = Font(name="Aptos", size=10, color=MUTED)
        ws[f"K{idx}"].font = Font(name="Aptos", size=10, bold=True, color=TEXT)
        ws[f"K{idx}"].fill = fill(NEUTRAL)
        apply_grid(ws, idx, idx, 10, 11)


def build_quote_summary(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 90
    style_title(ws, "Quote Summary", "Use this sheet to decide whether the quote is ready to send.", width=6)
    set_dimensions(ws, {"A": 28, "B": 18, "C": 18, "D": 18, "E": 20, "F": 24})

    cards = [
        (8, "Quote total", '=ROUND(SUM(\'Scope Inputs\'!H15:H29)*(1+IF(\'Scope Inputs\'!B9="yes",Assumptions!$B$8,0))-\'Scope Inputs\'!B11,2)', '$#,##0.00'),
        (9, "Deposit amount", '=ROUND(B8*Assumptions!$B$5,2)', '$#,##0.00'),
        (10, "Delivery window", '=IF(\'Scope Inputs\'!B9="yes",TEXT(Assumptions!$B$10,"0")&" business days",TEXT(Assumptions!$B$9,"0")&" business days")', '@'),
        (12, "Quote confidence state", '=IF(OR(\'Scope Inputs\'!B4="",\'Scope Inputs\'!B5="",\'Scope Inputs\'!B6<>"yes",\'Scope Inputs\'!B7<>"yes",\'Scope Inputs\'!B8<>"yes",COUNTIF(\'Scope Inputs\'!B15:B29,"<>")=0),"clarify_scope",IF(AND(\'Scope Inputs\'!B9="yes",\'Scope Inputs\'!B10<Assumptions!$B$10),"unrealistic_timeline",IF(B8<Assumptions!$B$6,"below_floor","send_ready")))', '@'),
        (13, "Recommended next step", '=IF(B12="clarify_scope","Ask for missing scope details before sending the quote",IF(B12="unrealistic_timeline","Reset delivery expectation before sending",IF(B12="below_floor","Raise scope, remove discount, or decline the job","Send the quote with deposit and timeline")))', '@'),
        (15, "Minimum profitable floor", '=Assumptions!$B$6', '$#,##0.00'),
        (16, "Quote margin buffer", '=ROUND(B8-B15,2)', '$#,##0.00'),
    ]
    for row, label, formula, numfmt in cards:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = formula
        ws[f"A{row}"].font = Font(name="Aptos", size=10, bold=True, color=MUTED)
        ws[f"B{row}"].font = Font(name="Aptos", size=12, bold=True, color=TEXT)
        ws[f"A{row}"].fill = fill(CARD)
        ws[f"B{row}"].fill = fill(OUTPUT)
        ws[f"A{row}"].alignment = Alignment(horizontal="left")
        ws[f"B{row}"].alignment = Alignment(horizontal="center" if row != 13 else "left", vertical="center", wrap_text=row == 13)
        ws[f"B{row}"].number_format = numfmt
        apply_grid(ws, row, row, 1, 2)
        ws.row_dimensions[row].height = 30 if row in (12, 13) else 24

    ws["D8"] = "What this means"
    ws["D8"].font = Font(name="Aptos", size=11, bold=True, color=TEXT)
    notes = [
        "send_ready: scope complete, timeline realistic, quote above floor",
        "clarify_scope: missing readiness input or no valid scope rows",
        "below_floor: total exists but job is too small at the current scope",
        "unrealistic_timeline: quote is feasible only if the delivery expectation changes",
    ]
    for idx, note in enumerate(notes, start=9):
        ws[f"D{idx}"] = note
        ws[f"D{idx}"].font = Font(name="Aptos", size=10, color=TEXT)
        ws.merge_cells(start_row=idx, start_column=4, end_row=idx, end_column=6)

    ws["D15"] = "Proof checklist"
    ws["D15"].font = Font(name="Aptos", size=11, bold=True, color=TEXT)
    proof = [
        "Totals respond to Assumptions and Scope Inputs.",
        "You should not send when the state is clarify_scope, below_floor, or unrealistic_timeline.",
        "Deposit and timeline update automatically from the current quote.",
    ]
    for idx, item in enumerate(proof, start=16):
        ws[f"D{idx}"] = "\u2022 " + item
        ws[f"D{idx}"].font = Font(name="Aptos", size=10, color=TEXT)
        ws.merge_cells(start_row=idx, start_column=4, end_row=idx, end_column=6)


def build_saved_quotes(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.sheet_view.zoomScale = 85
    style_title(ws, "Saved Quotes", "Use this log to keep prior quote snapshots visible without overwriting them.", width=10)
    set_dimensions(ws, {"A": 14, "B": 22, "C": 18, "D": 14, "E": 16, "F": 18, "G": 28, "H": 20, "I": 18, "J": 14})

    headers = [
        "Quote date",
        "Client",
        "Job type",
        "Quote total",
        "Deposit",
        "Delivery window",
        "State",
        "Next step",
        "Notes",
        "Version",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color=TEXT)
        cell.fill = fill(CARD)
        cell.alignment = Alignment(horizontal="center")
    apply_grid(ws, 4, 4, 1, 10)

    seeded = [
        ["2026-06-30", "Jordan Lee", "repair_bundle", 585, 175.5, "5 business days", "send_ready", "Send the quote with deposit and timeline", "Seeded standard quote", 1],
        ["2026-06-30", "Partial example", "drywall_patch", 188.5, 56.55, "5 business days", "clarify_scope", "Ask for missing scope details before sending the quote", "Shows blocked send state", 1],
    ]
    for row in range(5, 55):
        if row - 5 < len(seeded):
            values = seeded[row - 5]
            for col, value in enumerate(values, start=1):
                ws.cell(row, col, value)
        for col in range(1, 11):
            cell = ws.cell(row, col)
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(horizontal="left" if col in (2, 3, 7, 8, 9) else "center", vertical="center", wrap_text=col in (8, 9))
            if col <= 9:
                cell.fill = fill(INPUT if col <= 9 else SURFACE)
            else:
                cell.fill = fill(OUTPUT)
        apply_grid(ws, row, row, 1, 10)

    unlock_range(ws, "A5:I54")
    for row in range(5, 55):
        ws[f"D{row}"].number_format = '$#,##0.00'
        ws[f"E{row}"].number_format = '$#,##0.00'


def copy_sheet(source_ws, target_ws, value_ws=None):
    for row in source_ws.iter_rows():
        for cell in row:
            target = target_ws[cell.coordinate]
            target.value = value_ws[cell.coordinate].value if value_ws is not None else cell.value
            if cell.has_style:
                target._style = copy(cell._style)
            if cell.number_format:
                target.number_format = cell.number_format
            if cell.protection:
                target.protection = copy(cell.protection)
            if cell.alignment:
                target.alignment = copy(cell.alignment)
            if cell.font:
                target.font = copy(cell.font)
            if cell.fill:
                target.fill = copy(cell.fill)
            if cell.border:
                target.border = copy(cell.border)
    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dim.width
    for key, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key].height = dim.height
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
    target_ws.sheet_properties.tabColor = source_ws.sheet_properties.tabColor
    for merged in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged))


def recalc_roundtrip(source: Path, destination: Path):
    work_dir = PROCESS / "recalc"
    work_dir.mkdir(parents=True, exist_ok=True)
    ods_path = work_dir / f"{destination.stem}.ods"
    xlsx_path = work_dir / f"{destination.stem}.xlsx"
    for path in (ods_path, xlsx_path):
        if path.exists():
            path.unlink()
    subprocess.run([SOFFICE, "--headless", "--convert-to", "ods", "--outdir", str(work_dir), str(source)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    first_ods = work_dir / f"{source.stem}.ods"
    if first_ods != ods_path and first_ods.exists():
        first_ods.rename(ods_path)
    subprocess.run([SOFFICE, "--headless", "--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", str(work_dir), str(ods_path)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    generated_xlsx = work_dir / f"{ods_path.stem}.xlsx"
    if not generated_xlsx.exists():
        raise FileNotFoundError(generated_xlsx)
    shutil.copyfile(generated_xlsx, destination)


def export_surfaces(source_xlsx: Path):
    wb = load_workbook(source_xlsx)
    data_wb = load_workbook(source_xlsx, data_only=True)
    for sheet_name, slug in SURFACES:
        single = Workbook()
        target = single.active
        target.title = sheet_name
        copy_sheet(wb[sheet_name], target, data_wb[sheet_name])
        out_xlsx = SURFACE_XLSX_DIR / f"{slug}.xlsx"
        out_pdf = SURFACE_PDF_DIR / f"{slug}.pdf"
        out_png = SURFACE_THUMB_DIR / f"{slug}.pdf.png"
        render_png = RENDER_DIR / f"{slug}.png"
        single.save(out_xlsx)
        subprocess.run([SOFFICE, "--headless", "--convert-to", "pdf", "--outdir", str(SURFACE_PDF_DIR), str(out_xlsx)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.run(["sips", "-s", "format", "png", str(out_pdf), "--out", str(out_png)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        shutil.copyfile(out_png, render_png)


def write_delivery_files():
    (DELIVERY_DIR / "README-FIRST.md").write_text(
        "# Handyman Quote Readiness Planner\n\n"
        "Open `Handyman-Quote-Readiness-Planner.xlsx` first.\n\n"
        "1. Confirm your defaults in `Assumptions`.\n"
        "2. Complete the readiness inputs in `Scope Inputs`.\n"
        "3. Enter task rows only in yellow cells.\n"
        "4. Read `Quote Summary` and follow the next-step recommendation.\n\n"
        "Use `Saved Quotes` only to preserve snapshots. This product does not invoice, email clients, or schedule crews.\n",
        encoding="utf-8",
    )
    (DELIVERY_DIR / "Quick-Start-and-FAQ.md").write_text(
        "# Quick Start and FAQ\n\n"
        "## When is a quote ready to send?\n"
        "Only when `Quote Summary` shows `send_ready`.\n\n"
        "## Why did I get `clarify_scope`?\n"
        "One of the required readiness fields is incomplete or no valid scope row exists.\n\n"
        "## Why did I get `below_floor`?\n"
        "Your quote total is below the minimum profitable floor in `Assumptions`.\n\n"
        "## Why did I get `unrealistic_timeline`?\n"
        "The requested timing is faster than the rush turnaround rule.\n\n"
        "## Can I change the deposit percent?\n"
        "Yes. Update `Assumptions` and the summary recalculates.\n",
        encoding="utf-8",
    )


def main():
    for path in [
        PRODUCT_DIR,
        DELIVERY_DIR,
        SURFACE_XLSX_DIR,
        SURFACE_PDF_DIR,
        SURFACE_THUMB_DIR,
        RENDER_DIR,
        FORMULA_SOURCE.parent,
    ]:
        path.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    start_here = wb.create_sheet("Start Here")
    assumptions = wb.create_sheet("Assumptions")
    scope_inputs = wb.create_sheet("Scope Inputs")
    quote_summary = wb.create_sheet("Quote Summary")
    saved_quotes = wb.create_sheet("Saved Quotes")

    build_start_here(start_here)
    build_assumptions(assumptions)
    build_scope_inputs(scope_inputs)
    build_quote_summary(quote_summary)
    build_saved_quotes(saved_quotes)

    for ws in wb.worksheets:
        protect_sheet(ws)

    wb.security.workbookPassword = PROTECTION_PASSWORD
    wb.security.lockStructure = True
    wb.save(FORMULA_SOURCE)

    recalc_roundtrip(FORMULA_SOURCE, PRODUCT)
    shutil.copyfile(PRODUCT, IMPORT_COPY)
    export_surfaces(PRODUCT)
    write_delivery_files()

    print(PRODUCT)


if __name__ == "__main__":
    main()
