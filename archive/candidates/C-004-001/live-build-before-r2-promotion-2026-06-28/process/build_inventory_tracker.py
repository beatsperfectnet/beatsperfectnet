from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
from copy import copy
from datetime import date, timedelta
from pathlib import Path

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation


BASE = Path("/Users/andreyeremichev/beatsperfect/builds/C-004-001")
PRODUCT = BASE / "product" / "Inventory-Tracker-Studio.xlsx"
IMPORT_COPY = BASE / "product" / "Inventory-Tracker-Studio-google-sheets-import.xlsx"
PDF_CHECK = BASE / "process" / "pdf" / "Inventory-Tracker-Studio.pdf"
PREVIEW_DIR = BASE / "listing-assets" / "real-sheet-previews"
RENDER_DIR = BASE / "process" / "rendered"
DELIVERY_DIR = BASE / "delivery-assets"

BG = "F5F7F2"
SURFACE = "FFFFFF"
TEXT = "183153"
MUTED = "5A6472"
TEAL = "137A8A"
GREEN = "4E7C59"
CORAL = "D96C47"
GOLD = "D9A441"
INPUT = "FFF1BF"
OUTPUT = "DDEBF7"
SOFTGREEN = "DDEFD8"
CARD = "FAFBFC"
BORDER = "D8E1E8"

THIN = Side(style="thin", color=BORDER)
MED = Side(style="medium", color=TEAL)

PROTECTION_PASSWORD = "InventoryFlow"
YEAR = 2026
REVIEW_DATE = date(YEAR, 3, 15)
INVENTORY_CAPACITY = 100
PURCHASE_CAPACITY = 300
SALES_CAPACITY = 500
PRICE_CALC_CAPACITY = 100
PRICE_LIST_CAPACITY = 100

FAMILIES = [
    {"name": "Canvas Tote", "category": "Accessories", "supplier": "Northstar Wholesale", "base_cost": 0.95, "base_opening": 26, "base_reorder": 12, "note": "Reusable carry tote"},
    {"name": "Ceramic Mug", "category": "Drinkware", "supplier": "Harbor Supply", "base_cost": 2.10, "base_opening": 24, "base_reorder": 10, "note": "Gloss finish mug"},
    {"name": "A5 Notebook", "category": "Stationery", "supplier": "Paper Trail Co.", "base_cost": 1.25, "base_opening": 34, "base_reorder": 18, "note": "Sewn binding notebook"},
    {"name": "Sticker Pack", "category": "Stationery", "supplier": "Paper Trail Co.", "base_cost": 0.22, "base_opening": 42, "base_reorder": 28, "note": "Vinyl decal pack"},
    {"name": "Soy Candle", "category": "Home", "supplier": "Northstar Wholesale", "base_cost": 3.10, "base_opening": 16, "base_reorder": 8, "note": "8 oz amber jar"},
    {"name": "Keychain", "category": "Accessories", "supplier": "Luma Imports", "base_cost": 0.48, "base_opening": 28, "base_reorder": 14, "note": "Metal logo keychain"},
    {"name": "Water Bottle", "category": "Drinkware", "supplier": "Harbor Supply", "base_cost": 4.25, "base_opening": 18, "base_reorder": 9, "note": "Insulated steel bottle"},
    {"name": "Print Set", "category": "Wall Art", "supplier": "Studio Nine", "base_cost": 1.80, "base_opening": 22, "base_reorder": 11, "note": "2-piece art print set"},
    {"name": "Phone Grip", "category": "Accessories", "supplier": "Luma Imports", "base_cost": 0.62, "base_opening": 52, "base_reorder": 20, "note": "Silicone grip stand"},
    {"name": "Crew Socks", "category": "Apparel", "supplier": "Northstar Wholesale", "base_cost": 1.70, "base_opening": 30, "base_reorder": 15, "note": "One-size crew socks"},
    {"name": "Softcover Journal", "category": "Stationery", "supplier": "Paper Trail Co.", "base_cost": 2.30, "base_opening": 26, "base_reorder": 13, "note": "Dot grid journal"},
    {"name": "Enamel Pin", "category": "Accessories", "supplier": "Studio Nine", "base_cost": 0.85, "base_opening": 20, "base_reorder": 10, "note": "Hard enamel pin"},
    {"name": "Lunch Box", "category": "Home", "supplier": "Harbor Supply", "base_cost": 3.40, "base_opening": 14, "base_reorder": 8, "note": "Snap lid lunch box"},
    {"name": "Desk Mat", "category": "Accessories", "supplier": "Northstar Wholesale", "base_cost": 5.20, "base_opening": 12, "base_reorder": 6, "note": "Large desk mat"},
    {"name": "Gift Tag Set", "category": "Stationery", "supplier": "Paper Trail Co.", "base_cost": 0.58, "base_opening": 38, "base_reorder": 22, "note": "Mixed tag pack"},
    {"name": "Washi Tape", "category": "Stationery", "supplier": "Paper Trail Co.", "base_cost": 0.72, "base_opening": 36, "base_reorder": 18, "note": "Decorative tape roll"},
    {"name": "Tote Pouch", "category": "Accessories", "supplier": "Luma Imports", "base_cost": 1.10, "base_opening": 24, "base_reorder": 12, "note": "Zip pouch insert"},
    {"name": "Metal Bookmark", "category": "Stationery", "supplier": "Studio Nine", "base_cost": 0.40, "base_opening": 46, "base_reorder": 20, "note": "Laser-cut bookmark"},
    {"name": "Wall Print", "category": "Wall Art", "supplier": "Studio Nine", "base_cost": 2.90, "base_opening": 18, "base_reorder": 9, "note": "Frameless print"},
    {"name": "Tea Tin", "category": "Home", "supplier": "Harbor Supply", "base_cost": 1.95, "base_opening": 16, "base_reorder": 8, "note": "Loose leaf tea tin"},
]

MODIFIERS = [
    "Sand",
    "Olive",
    "Slate",
    "Coral",
    "Natural",
]

CHANNELS = [
    "Etsy",
    "Shop",
    "Market",
    "Retail",
    "Wholesale",
]


def inventory_code(idx: int) -> str:
    return f"SKU-{idx:03d}"


def currency(value: float) -> float:
    return round(value, 2)


def build_inventory_items():
    rows = []
    for idx in range(1, INVENTORY_CAPACITY + 1):
        family = FAMILIES[(idx - 1) // len(MODIFIERS)]
        modifier = MODIFIERS[(idx - 1) % len(MODIFIERS)]
        variant_index = (idx - 1) // len(FAMILIES) + 1
        if idx <= 4:
            opening = 2 + ((idx - 1) % 2)
        elif idx <= 10:
            opening = 10 + ((idx - 1) % 3)
        else:
            opening = family["base_opening"] + ((idx - 1) % 6) * 2 + ((idx - 1) // 10)
        reorder = family["base_reorder"] + ((idx - 1) % 4)
        unit_cost = currency(family["base_cost"] + ((idx - 1) % 5) * 0.17 + ((idx - 1) // 25) * 0.08)
        rows.append(
            (
                inventory_code(idx),
                f"{family['name']} - {modifier}",
                family["category"],
                family["supplier"],
                opening,
                0,
                unit_cost,
                reorder,
                f"{family['note']} / batch {variant_index}",
            )
        )
    return rows


inventory_items = build_inventory_items()


def build_purchases():
    rows = []
    start_date = date(YEAR, 1, 3)
    for idx in range(1, PURCHASE_CAPACITY + 1):
        item = inventory_items[(idx - 1) % INVENTORY_CAPACITY]
        date_offset = (idx - 1) // 6
        qty = 2 + ((idx - 1) % 5)
        unit_cost = item[6]
        shipping = currency(1.80 + ((idx - 1) % 7) * 0.35 + ((idx - 1) // 75) * 0.12)
        rows.append(
            (
                start_date + timedelta(days=date_offset),
                item[0],
                item[3],
                qty,
                unit_cost,
                shipping,
                f"Restock wave {((idx - 1) // INVENTORY_CAPACITY) + 1}",
            )
        )
    return rows


purchases = build_purchases()


def build_sales():
    rows = []
    start_date = date(YEAR, 1, 4)
    for idx in range(1, SALES_CAPACITY + 1):
        item = inventory_items[(idx - 1) % INVENTORY_CAPACITY]
        date_offset = (idx - 1) // 5
        sku_position = ((idx - 1) % INVENTORY_CAPACITY) + 1
        if sku_position <= 4:
            qty = 4 + ((idx - 1) % 2)
        elif sku_position <= 10:
            qty = 3 + ((idx - 1) % 2)
        else:
            qty = 1 + ((idx - 1) % 3)
        unit_price = currency(item[6] * (2.8 + ((idx - 1) % 6) * 0.08))
        fees = currency(unit_price * 0.13)
        rows.append(
            (
                start_date + timedelta(days=date_offset),
                item[0],
                CHANNELS[(idx - 1) % len(CHANNELS)],
                qty,
                unit_price,
                fees,
                f"Order batch {((idx - 1) // INVENTORY_CAPACITY) + 1}",
            )
        )
    return rows


sales = build_sales()


def build_price_calc_rows():
    rows = []
    for idx, item in enumerate(inventory_items, start=1):
        rows.append(
            (
                item[1],
                item[0],
                item[3],
                item[6],
                currency(0.18 + ((idx - 1) % 4) * 0.07),
                currency(0.44 + ((idx - 1) % 5) * 0.09),
                0.13,
                0.10,
                0.42 + ((idx - 1) % 3) * 0.03,
                f"Pricing scenario {idx}",
            )
        )
    return rows


price_calc_rows = build_price_calc_rows()


def fill(rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=rgb)


def border(left=False, right=False, top=False, bottom=False) -> Border:
    return Border(
        left=MED if left else THIN,
        right=MED if right else THIN,
        top=MED if top else THIN,
        bottom=MED if bottom else THIN,
    )


def apply_cell(cell, *, fill_rgb=None, font=None, align=None, bold=False, border_spec=None):
    if fill_rgb:
        cell.fill = fill(fill_rgb)
    if font:
        cell.font = copy(font)
    elif bold:
        cell.font = Font(name="Inter", size=9, bold=True, color=TEXT)
    if align:
        cell.alignment = copy(align)
    if border_spec:
        cell.border = border(**border_spec)


def title_bar(ws, title, subtitle, end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].font = Font(name="Inter", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = fill(TEAL)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Inter", size=10, italic=True, color=MUTED)
    ws["A2"].fill = fill(BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22


def table_header(ws, row, headers, fill_rgb=CARD):
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row, c, header)
        cell.font = Font(name="Inter", size=9, bold=True, color=TEXT)
        cell.fill = fill(fill_rgb)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border(top=True, bottom=True)


def add_sheet_basics(ws, zoom=90):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    ws.freeze_panes = "A4"


def add_card(ws, start_col, row, label, formula_or_value, fill_rgb=INPUT, width=2):
    ws.cell(row, start_col, label)
    ws.cell(row, start_col).font = Font(name="Inter", size=8, bold=True, color=MUTED)
    ws.cell(row, start_col).fill = fill(CARD)
    ws.cell(row, start_col).alignment = Alignment(horizontal="left")
    ws.cell(row, start_col + 1, formula_or_value)
    ws.cell(row, start_col + 1).font = Font(name="IBM Plex Mono", size=13, bold=True, color=TEXT)
    ws.cell(row, start_col + 1).fill = fill(fill_rgb)
    ws.cell(row, start_col + 1).alignment = Alignment(horizontal="center")
    ws.cell(row, start_col).border = border(left=True, top=True, bottom=True)
    ws.cell(row, start_col + 1).border = border(right=True, top=True, bottom=True)
    ws.row_dimensions[row].height = 24


def unlock_range(ws, range_ref: str):
    for row in ws[range_ref]:
        for cell in row:
            cell.protection = Protection(locked=False)


def build_start_here(ws):
    add_sheet_basics(ws, zoom=95)
    title_bar(ws, "Inventory Tracker Studio", "Start with Setup, then use Inventory to decide what to reorder, when, and how much.", 8)
    ws["A4"] = "What to do first"
    ws["A4"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    steps = [
        "Set your review date, lead time, safety stock days, pack size, fees, and target margin in Setup.",
        "Edit the yellow cells on Inventory for your stock, suppliers, and reorder points.",
        "Enter purchases and sales in the log tabs; stock and profit will update automatically.",
        "Use Inventory to see reorder action, reorder-by date, projected stockout, days of cover, and suggested order quantity.",
        "Use Price Calculator to test sale prices before you list items.",
    ]
    for i, step in enumerate(steps, start=5):
        ws.cell(i, 1, f"{i-4}. {step}")
        ws.cell(i, 1).font = Font(name="Inter", size=10, color=TEXT)
    ws["D4"] = "Editable areas"
    ws["D4"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    editable = [
        "Setup input cells",
        "Inventory opening stock, reorder point, and unit cost",
        "Purchase log rows",
        "Sales log rows",
        "Price Calculator inputs",
    ]
    for i, text in enumerate(editable, start=5):
        ws.cell(i, 4, text)
        ws.cell(i, 4).font = Font(name="Inter", size=10, color=TEXT)
    ws["D11"] = "Leave alone"
    ws["D11"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    locked = [
        "Dashboard formulas and charts",
        "Inventory status formulas",
        "Sales profit formulas",
        "Price List snapshot formulas",
    ]
    for i, text in enumerate(locked, start=12):
        ws.cell(i, 4, text)
        ws.cell(i, 4).font = Font(name="Inter", size=10, color=TEXT)
    ws["A13"] = "Recovery"
    ws["A13"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    ws["A14"] = "Keep a clean copy before heavy edits. If you damage formulas, start again from a fresh duplicate."
    ws["A14"].font = Font(name="Inter", size=9, color=MUTED)
    ws["A16"] = "What this is not"
    ws["A16"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    noes = [
        "Not a subscription app",
        "Not bank-connected",
        "Not tax or accounting advice",
        "Not a live inventory service",
    ]
    for i, text in enumerate(noes, start=17):
        ws.cell(i, 1, f"• {text}")
        ws.cell(i, 1).font = Font(name="Inter", size=10, color=TEXT)
    ws["D17"] = "Seeded capacity"
    ws["D17"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    ws["D18"] = "100 inventory rows, 300 purchase rows, 500 sales rows, 100 price calculator rows, and 100 price list rows."
    ws["D18"].alignment = Alignment(wrap_text=True)
    ws["D18"].font = Font(name="Inter", size=9, color=TEXT)
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 40


def build_setup(ws):
    add_sheet_basics(ws)
    title_bar(ws, "Setup", "Choose the defaults that drive replenishment, stock, and pricing formulas.", 6)
    settings = [
        ("Currency symbol", "$"),
        ("Review date", REVIEW_DATE),
        ("Sales lookback days", 30),
        ("Sales tax", 0.10),
        ("Marketplace fee", 0.13),
        ("Target margin", 0.45),
        ("Default reorder point", 10),
        ("Lead time days", 14),
        ("Safety stock days", 7),
        ("Default pack size", 6),
        ("Support note", "Use the workbook for stock, purchases, sales, and pricing."),
    ]
    for i, (label, value) in enumerate(settings, start=4):
        ws.cell(i, 1, label)
        ws.cell(i, 1).font = Font(name="Inter", size=9, bold=True, color=MUTED)
        ws.cell(i, 2, value)
        ws.cell(i, 2).font = Font(name="IBM Plex Mono", size=10, bold=True, color=TEXT)
        ws.cell(i, 2).fill = fill(INPUT)
        ws.cell(i, 1).border = border(left=True, top=True, bottom=True)
        ws.cell(i, 2).border = border(right=True, top=True, bottom=True)
        if label == "Review date":
            ws.cell(i, 2).number_format = "yyyy-mm-dd"
    ws["D4"] = "Categories"
    ws["D4"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    categories = [
        "Accessories",
        "Drinkware",
        "Stationery",
        "Home",
        "Wall Art",
        "Apparel",
    ]
    ws["D5"] = ", ".join(categories)
    ws["D5"].alignment = Alignment(wrap_text=True)
    ws["D5"].font = Font(name="Inter", size=9, color=TEXT)
    ws["D7"] = "Editable cells"
    ws["D7"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    ws["D8"] = "Yellow input cells in Setup, Inventory, Purchases, Sales, and Price Calculator."
    ws["D8"].font = Font(name="Inter", size=9, color=TEXT)
    ws["D13"] = "Workbook capacity"
    ws["D13"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    ws["D14"] = "The delivered workbook ships with 100 inventory rows, 300 purchase rows, 500 sales rows, 100 calculator rows, and 100 price-list rows."
    ws["D14"].alignment = Alignment(wrap_text=True)
    ws["D14"].font = Font(name="Inter", size=9, color=TEXT)
    ws["D16"] = "Unsupported"
    ws["D16"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    ws["D17"] = "Live bank sync, tax filing, and custom bookkeeping."
    ws["D17"].font = Font(name="Inter", size=9, color=TEXT)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 62


def build_inventory(ws):
    add_sheet_basics(ws)
    title_bar(ws, "Inventory", "Decide what to reorder, when to reorder it, and how much to buy.", 21)
    headers = [
        "SKU", "Item", "Category", "Supplier", "Opening Stock", "Purchases", "Sales",
        "On Hand", "Reorder Point", "Unit Cost", "Inventory Value", "Sales Velocity/Day",
        "Days Cover", "Projected Stockout", "Lead Time Used", "Safety Stock",
        "Suggested Order Qty", "Reorder By", "Reorder Action", "Status", "Notes"
    ]
    table_header(ws, 4, headers)
    first_row = 5
    for idx, row in enumerate(inventory_items, start=first_row):
        sku, item, category, supplier, opening, purchase_qty, unit_cost, reorder, note = row
        ws.cell(idx, 1, sku)
        ws.cell(idx, 2, item)
        ws.cell(idx, 3, category)
        ws.cell(idx, 4, supplier)
        ws.cell(idx, 5, opening)
        ws.cell(idx, 6, f'=SUMIFS(Purchases!$D$5:$D$304,Purchases!$B$5:$B$304,$A{idx})')
        ws.cell(idx, 7, f'=SUMIFS(Sales!$D$5:$D$504,Sales!$B$5:$B$504,$A{idx})')
        ws.cell(idx, 8, f"=E{idx}+F{idx}-G{idx}")
        ws.cell(idx, 9, reorder)
        ws.cell(idx, 10, unit_cost)
        ws.cell(idx, 11, f"=H{idx}*J{idx}")
        ws.cell(idx, 12, f'=IFERROR(SUMIFS(Sales!$D$5:$D$504,Sales!$B$5:$B$504,$A{idx},Sales!$A$5:$A$504,">="&Setup!$B$5-Setup!$B$6,Sales!$A$5:$A$504,"<="&Setup!$B$5)/Setup!$B$6,0)')
        ws.cell(idx, 13, f'=IF(L{idx}<=0,999,H{idx}/L{idx})')
        ws.cell(idx, 14, f'=IF(L{idx}<=0,"No recent sales",Setup!$B$5+M{idx})')
        ws.cell(idx, 15, "=Setup!$B$11")
        ws.cell(idx, 16, f"=ROUNDUP(L{idx}*Setup!$B$12,0)")
        ws.cell(idx, 17, f"=IF(MAX(0,ROUNDUP((L{idx}*(O{idx}+Setup!$B$12))-H{idx},0))=0,0,CEILING(MAX(0,ROUNDUP((L{idx}*(O{idx}+Setup!$B$12))-H{idx},0)),Setup!$B$13))")
        ws.cell(idx, 18, f'=IF(L{idx}<=0,"No sales signal",N{idx}-O{idx})')
        ws.cell(idx, 19, f'=IF(H{idx}<=0,"Order now",IF(R{idx}="No sales signal","No action",IF(R{idx}<=Setup!$B$5,"Order now",IF(R{idx}<=Setup!$B$5+7,"Order this week",IF(H{idx}<=I{idx},"Plan reorder","No action")))))')
        ws.cell(idx, 20, f'=IF(H{idx}<=0,"Out of stock",IF(H{idx}<=I{idx},"Reorder soon","Healthy"))')
        ws.cell(idx, 21, note)
        for c in range(1, 22):
            cell = ws.cell(idx, c)
            cell.font = Font(name="Inter", size=9, color=TEXT)
            cell.alignment = Alignment(horizontal="left" if c in (1, 2, 3, 4, 19, 20, 21) else "center")
            cell.fill = fill(INPUT if c in (5, 9, 10, 21) else OUTPUT if c in (6, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20) else SURFACE)
            cell.border = border()
        for c in (5, 6, 7, 8, 9, 15, 16, 17):
            ws.cell(idx, c).number_format = '0'
        ws.cell(idx, 10).number_format = '$#,##0.00'
        ws.cell(idx, 11).number_format = '$#,##0.00'
        ws.cell(idx, 12).number_format = '0.00'
        ws.cell(idx, 13).number_format = '0.0'
        ws.cell(idx, 14).number_format = 'yyyy-mm-dd'
        ws.cell(idx, 18).number_format = 'yyyy-mm-dd'
    total_row = first_row + len(inventory_items)
    ws.cell(total_row, 1, "Totals")
    ws.cell(total_row, 5, f"=SUM(E{first_row}:E{total_row-1})")
    ws.cell(total_row, 6, f"=SUM(F{first_row}:F{total_row-1})")
    ws.cell(total_row, 7, f"=SUM(G{first_row}:G{total_row-1})")
    ws.cell(total_row, 8, f"=SUM(H{first_row}:H{total_row-1})")
    ws.cell(total_row, 11, f"=SUM(K{first_row}:K{total_row-1})")
    ws.cell(total_row, 17, f"=SUM(Q{first_row}:Q{total_row-1})")
    for c in range(1, 22):
        cell = ws.cell(total_row, c)
        cell.font = Font(name="Inter", size=9, bold=True, color=TEXT)
        cell.fill = fill(CARD)
        cell.border = border(bottom=True)
    for c in (5, 6, 7, 8, 17):
        ws.cell(total_row, c).number_format = '0'
    ws.cell(total_row, 11).number_format = '$#,##0.00'
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 17
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["N"].width = 18
    ws.column_dimensions["O"].width = 14
    ws.column_dimensions["P"].width = 13
    ws.column_dimensions["Q"].width = 18
    ws.column_dimensions["R"].width = 14
    ws.column_dimensions["S"].width = 16
    ws.column_dimensions["T"].width = 14
    ws.column_dimensions["U"].width = 24


def build_purchases(ws):
    add_sheet_basics(ws)
    title_bar(ws, "Purchases", "Log restocks and purchase costs. Quantity drives Inventory on-hand and total cost drives the dashboard.", 10)
    headers = ["Date", "SKU", "Supplier", "Qty", "Unit Cost", "Shipping", "Total Cost", "Notes"]
    table_header(ws, 4, headers)
    for idx, row in enumerate(purchases, start=5):
        d, sku, supplier, qty, unit_cost, shipping, note = row
        ws.cell(idx, 1, d)
        ws.cell(idx, 2, sku)
        ws.cell(idx, 3, supplier)
        ws.cell(idx, 4, qty)
        ws.cell(idx, 5, unit_cost)
        ws.cell(idx, 6, shipping)
        ws.cell(idx, 7, f"=D{idx}*E{idx}+F{idx}")
        ws.cell(idx, 8, note)
        for c in range(1, 9):
            cell = ws.cell(idx, c)
            cell.font = Font(name="Inter", size=9, color=TEXT)
            cell.alignment = Alignment(horizontal="left" if c in (1, 2, 3, 8) else "center")
            cell.fill = fill(INPUT if c in (1, 2, 3, 4, 5, 6, 8) else OUTPUT)
            cell.border = border()
        ws.cell(idx, 4).number_format = '0'
        ws.cell(idx, 1).number_format = "yyyy-mm-dd"
        ws.cell(idx, 5).number_format = '$#,##0.00'
        ws.cell(idx, 6).number_format = '$#,##0.00'
        ws.cell(idx, 7).number_format = '$#,##0.00'
    total_row = 5 + len(purchases)
    ws.cell(total_row, 1, "Totals")
    ws.cell(total_row, 4, f"=SUM(D5:D{total_row - 1})")
    ws.cell(total_row, 7, f"=SUM(G5:G{total_row - 1})")
    for c in range(1, 9):
        cell = ws.cell(total_row, c)
        cell.font = Font(name="Inter", size=9, bold=True, color=TEXT)
        cell.fill = fill(CARD)
        cell.border = border(bottom=True)
    ws.cell(total_row, 4).number_format = '0'
    ws.cell(total_row, 7).number_format = '$#,##0.00'
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 24


def build_sales(ws):
    add_sheet_basics(ws)
    title_bar(ws, "Sales", "Log sales orders and see revenue, fees, cost of goods sold, and gross profit.", 10)
    headers = ["Date", "SKU", "Channel", "Qty", "Unit Price", "Fees", "Revenue", "COGS", "Gross Profit", "Notes"]
    table_header(ws, 4, headers)
    for idx, row in enumerate(sales, start=5):
        d, sku, channel, qty, unit_price, fees, note = row
        ws.cell(idx, 1, d)
        ws.cell(idx, 2, sku)
        ws.cell(idx, 3, channel)
        ws.cell(idx, 4, qty)
        ws.cell(idx, 5, unit_price)
        ws.cell(idx, 6, fees)
        ws.cell(idx, 7, f"=D{idx}*E{idx}")
        ws.cell(idx, 8, f'=D{idx}*VLOOKUP(B{idx},Inventory!$A$5:$J$104,10,FALSE)')
        ws.cell(idx, 9, f"=G{idx}-F{idx}-H{idx}")
        ws.cell(idx, 10, note)
        for c in range(1, 11):
            cell = ws.cell(idx, c)
            cell.font = Font(name="Inter", size=9, color=TEXT)
            cell.alignment = Alignment(horizontal="left" if c in (1, 2, 3, 10) else "center")
            cell.fill = fill(INPUT if c in (1, 2, 3, 4, 5, 6, 10) else OUTPUT)
            cell.border = border()
        ws.cell(idx, 4).number_format = '0'
        ws.cell(idx, 1).number_format = "yyyy-mm-dd"
        ws.cell(idx, 5).number_format = '$#,##0.00'
        ws.cell(idx, 6).number_format = '$#,##0.00'
        ws.cell(idx, 7).number_format = '$#,##0.00'
        ws.cell(idx, 8).number_format = '$#,##0.00'
        ws.cell(idx, 9).number_format = '$#,##0.00'
    total_row = 5 + len(sales)
    ws.cell(total_row, 1, "Totals")
    ws.cell(total_row, 4, f"=SUM(D5:D{total_row - 1})")
    ws.cell(total_row, 7, f"=SUM(G5:G{total_row - 1})")
    ws.cell(total_row, 8, f"=SUM(H5:H{total_row - 1})")
    ws.cell(total_row, 9, f"=SUM(I5:I{total_row - 1})")
    for c in range(1, 11):
        cell = ws.cell(total_row, c)
        cell.font = Font(name="Inter", size=9, bold=True, color=TEXT)
        cell.fill = fill(CARD)
        cell.border = border(bottom=True)
    for c in (4,):
        ws.cell(total_row, c).number_format = '0'
    for c in (7, 8, 9):
        ws.cell(total_row, c).number_format = '$#,##0.00'
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 24


def build_price_calculator(ws):
    add_sheet_basics(ws)
    title_bar(ws, "Price Calculator", "Test sale price, expected profit, and margin before you list a product.", 13)
    headers = ["Product", "SKU", "COGS", "Packaging", "Labor", "Overhead", "Fee %", "Tax %", "Target Margin", "Suggested Price", "Profit", "Margin", "Notes"]
    table_header(ws, 4, headers)
    for idx, row in enumerate(price_calc_rows, start=5):
        product, sku, supplier, cost, pack, labor, fee, tax, margin, note = row
        ws.cell(idx, 1, product)
        ws.cell(idx, 2, sku)
        ws.cell(idx, 3, f'=VLOOKUP(B{idx},Inventory!$A$5:$J$104,10,FALSE)')
        ws.cell(idx, 4, pack)
        ws.cell(idx, 5, labor)
        ws.cell(idx, 6, 0.25)
        ws.cell(idx, 7, fee)
        ws.cell(idx, 8, tax)
        ws.cell(idx, 9, margin)
        ws.cell(idx, 10, f'=IF(A{idx}="","",ROUND((C{idx}+D{idx}+E{idx}+F{idx})/(1-(G{idx}+H{idx}+I{idx})),2))')
        ws.cell(idx, 11, f'=IF(J{idx}="","",J{idx}-(C{idx}+D{idx}+E{idx}+F{idx})-(J{idx}*G{idx})-(J{idx}*H{idx}))')
        ws.cell(idx, 12, f'=IF(J{idx}="","",IF(J{idx}=0,"",K{idx}/J{idx}))')
        ws.cell(idx, 13, note)
        for c in range(1, 14):
            cell = ws.cell(idx, c)
            cell.font = Font(name="Inter", size=9, color=TEXT)
            cell.alignment = Alignment(horizontal="left" if c in (1, 2, 13) else "center")
            cell.fill = fill(INPUT if c in (1, 2, 4, 5, 6, 7, 8, 9, 13) else OUTPUT)
            cell.border = border()
        for c in (3, 4, 5, 6, 10, 11):
            ws.cell(idx, c).number_format = '$#,##0.00'
        for c in (7, 8, 9, 12):
            ws.cell(idx, c).number_format = '0%'
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 10
    ws.column_dimensions["M"].width = 20


def build_price_list(ws):
    add_sheet_basics(ws)
    title_bar(ws, "Price List", "Save approved price snapshots here. The top row mirrors the calculator output for quick reuse.", 9)
    headers = ["Date", "SKU", "Product", "Cost", "Suggested Price", "Profit", "Margin", "Status", "Notes"]
    table_header(ws, 4, headers)
    for idx, row in enumerate(price_calc_rows, start=5):
        product, sku, supplier, cost, pack, labor, fee, tax, margin, note = row
        ws.cell(idx, 1, date(YEAR, 1, 20) + timedelta(days=idx - 5))
        ws.cell(idx, 2, sku)
        ws.cell(idx, 3, product)
        ws.cell(idx, 4, f'=VLOOKUP(B{idx},Inventory!$A$5:$J$104,10,FALSE)')
        ws.cell(idx, 5, f"='Price Calculator'!J{idx}")
        ws.cell(idx, 6, f"='Price Calculator'!K{idx}")
        ws.cell(idx, 7, f"='Price Calculator'!L{idx}")
        ws.cell(idx, 8, f'=IF(G{idx}>=0.45,"Good margin","Review")')
        ws.cell(idx, 9, note)
        for c in range(1, 10):
            cell = ws.cell(idx, c)
            cell.font = Font(name="Inter", size=9, color=TEXT)
            cell.alignment = Alignment(horizontal="left" if c in (1, 2, 3, 9) else "center")
            cell.fill = fill(INPUT if c in (1, 2, 3, 9) else OUTPUT)
            cell.border = border()
        ws.cell(idx, 4).number_format = '$#,##0.00'
        ws.cell(idx, 5).number_format = '$#,##0.00'
        ws.cell(idx, 6).number_format = '$#,##0.00'
        ws.cell(idx, 7).number_format = '0%'
        ws.cell(idx, 1).number_format = "yyyy-mm-dd"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 20


def build_dashboard(ws):
    add_sheet_basics(ws, zoom=95)
    title_bar(ws, "Dashboard", "The replenishment decision surface: what to reorder, when, and how much.", 16)
    # summary cards
    cards = [
        ("B4", "Inventory rows", "=COUNTA(Inventory!A5:A104)", OUTPUT),
        ("D4", "Order now", '=COUNTIF(Inventory!S5:S104,"Order now")', CORAL),
        ("F4", "Order qty", '=SUM(Inventory!Q5:Q104)', GOLD),
        ("H4", "Inventory value", '=SUM(Inventory!K5:K104)', SOFTGREEN),
        ("J4", "Purchase spend", '=SUM(Purchases!G5:G304)', INPUT),
        ("L4", "Sales revenue", '=SUM(Sales!G5:G504)', OUTPUT),
        ("N4", "Gross profit", '=SUM(Sales!I5:I504)', SOFTGREEN),
    ]
    for pos, label, formula, color in cards:
        row = ws[pos].row
        col = ws[pos].column
        ws.cell(row, col, label)
        ws.cell(row, col).font = Font(name="Inter", size=8, bold=True, color=MUTED)
        ws.cell(row + 1, col, formula)
        ws.cell(row + 1, col).font = Font(name="IBM Plex Mono", size=13, bold=True, color=TEXT)
        ws.cell(row + 1, col).fill = fill(color)
        ws.cell(row, col).fill = fill(CARD)
        ws.cell(row, col).border = border()
        ws.cell(row + 1, col).border = border()
        ws.cell(row + 1, col).number_format = '$#,##0.00' if label not in {"Inventory rows", "Order now", "Order qty"} else '0'
    ws["A8"] = "Reorder action mix"
    ws["A8"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    status_rows = [
        ("Order now", '=COUNTIF(Inventory!S5:S104,"Order now")'),
        ("Order this week", '=COUNTIF(Inventory!S5:S104,"Order this week")'),
        ("Plan reorder", '=COUNTIF(Inventory!S5:S104,"Plan reorder")'),
        ("No action", '=COUNTIF(Inventory!S5:S104,"No action")'),
    ]
    table_header(ws, 9, ["Status", "Count"])
    for i, (name, formula) in enumerate(status_rows, start=10):
        ws.cell(i, 1, name)
        ws.cell(i, 2, formula)
        ws.cell(i, 1).font = Font(name="Inter", size=9, color=TEXT)
        ws.cell(i, 2).font = Font(name="IBM Plex Mono", size=9, bold=True, color=TEXT)
        ws.cell(i, 1).fill = fill(SURFACE)
        ws.cell(i, 2).fill = fill(OUTPUT)
        ws.cell(i, 1).border = border()
        ws.cell(i, 2).border = border()
    bar = BarChart()
    bar.title = "Inventory value by item"
    bar.y_axis.title = "USD"
    data = Reference(ws.parent["Inventory"], min_col=11, min_row=4, max_row=104)
    cats = Reference(ws.parent["Inventory"], min_col=2, min_row=5, max_row=104)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 7.0
    bar.width = 14.5
    bar.legend = None
    ws.add_chart(bar, "G8")
    pie = PieChart()
    pie.title = "Inventory status"
    data2 = Reference(ws, min_col=2, min_row=9, max_row=13)
    labels2 = Reference(ws, min_col=1, min_row=10, max_row=13)
    pie.add_data(data2, titles_from_data=True)
    pie.set_categories(labels2)
    pie.height = 7.0
    pie.width = 10.5
    pie.dataLabels = None
    ws.add_chart(pie, "N8")
    ws["A16"] = "Top notes"
    ws["A16"].font = Font(name="Inter", size=11, bold=True, color=TEXT)
    notes = [
        "Use Setup first, then Inventory.",
        "Inventory outputs reorder action, reorder-by date, projected stockout date, days cover, suggested order qty, velocity, lead time, and safety stock.",
        "Seeded capacity: 100 inventory rows, 300 purchase rows, 500 sales rows, 100 calculator rows, 100 price-list rows.",
        "Purchases raise stock; Sales lower stock.",
        "Price Calculator helps you test margins before listing.",
        "Price List stores the approved price snapshot.",
    ]
    for i, text in enumerate(notes, start=17):
        ws.cell(i, 1, f"• {text}")
        ws.cell(i, 1).font = Font(name="Inter", size=10, color=TEXT)
    ws.column_dimensions["A"].width = 32
    for col in "BCDEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 12
    for col in ["B", "D", "F", "H", "J", "L", "N"]:
        ws.column_dimensions[col].width = 16


def build_help(ws):
    add_sheet_basics(ws, zoom=95)
    title_bar(ws, "Help", "Compatibility, limits, and support scope. Keep this sheet visible if you share the file.", 8)
    sections = {
        "Compatible with": [
            "Excel-compatible spreadsheet editors",
            "Google Sheets standard-formula import",
        ],
        "You can change": [
            "Yellow input cells in Setup",
            "Yellow input cells on Inventory, Purchases, Sales, and Price Calculator",
        ],
        "Leave alone": [
            "Dashboard formulas and charts",
            "Inventory status formulas",
            "Sales profit formulas",
            "Price List snapshot formulas",
        ],
        "Support scope": [
            "File access, setup guidance, and workbook questions",
            "Not bookkeeping, tax advice, or live account connections",
        ],
        "Limits": [
            "This workbook does not connect to banks or live inventory services.",
            "Imported protection settings can vary by editor or account.",
        ],
    }
    row = 4
    for title, bullets in sections.items():
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(name="Inter", size=11, bold=True, color=TEXT)
        for bullet in bullets:
            row += 1
            ws.cell(row, 1, f"• {bullet}")
            ws.cell(row, 1).font = Font(name="Inter", size=10, color=TEXT)
        row += 2
    ws.column_dimensions["A"].width = 94


def build_lists(ws):
    ws.sheet_state = "hidden"
    ws["A1"] = "SKUs"
    ws["B1"] = "Categories"
    for idx, item in enumerate(inventory_items, start=2):
        ws.cell(idx, 1, item[0])
    for idx, category in enumerate(["Accessories", "Drinkware", "Stationery", "Home", "Wall Art", "Apparel"], start=2):
        ws.cell(idx, 2, category)


def configure_validation(wb):
    wb["Inventory"].add_data_validation(DataValidation(type="list", formula1="=Lists!$B$2:$B$7", allow_blank=True))
    wb["Inventory"].data_validations.dataValidation[-1].add("C5:C104")
    for ws_name in ["Purchases", "Sales"]:
        ws = wb[ws_name]
        dv = DataValidation(type="list", formula1="=Lists!$A$2:$A$101", allow_blank=True)
        ws.add_data_validation(dv)
        limit_row = 304 if ws_name == "Purchases" else 504
        dv.add(f"B5:B{limit_row}")
    fee_dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=False)
    wb["Setup"].add_data_validation(fee_dv)
    fee_dv.add("B7:B9")
    # price calculator and list use sku selection validation
    for ws_name in ["Price Calculator", "Price List"]:
        ws = wb[ws_name]
        dv = DataValidation(type="list", formula1="=Lists!$A$2:$A$101", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("B5:B104")


def protect_surfaces(wb):
    wb.security.lockStructure = True
    wb.security.set_workbook_password(PROTECTION_PASSWORD)
    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.set_password(PROTECTION_PASSWORD)
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = True
    unlock_range(wb["Setup"], "B4:B14")
    unlock_range(wb["Inventory"], "E5:E104")
    unlock_range(wb["Inventory"], "I5:J104")
    unlock_range(wb["Inventory"], "U5:U104")
    unlock_range(wb["Purchases"], "A5:H304")
    unlock_range(wb["Sales"], "A5:J504")
    unlock_range(wb["Price Calculator"], "A5:M104")
    unlock_range(wb["Price List"], "A5:I104")


def add_comments(wb):
    wb["Start Here"]["A4"].comment = Comment("This workbook is designed to be edited from the yellow cells only.", "BeatsPerfect")
    wb["Help"]["A4"].comment = Comment("Use this sheet for support and limits.", "BeatsPerfect")


def workbook_metadata(wb):
    wb.properties.creator = "BeatsPerfect"
    wb.properties.title = "Inventory Tracker Studio"
    wb.properties.subject = "Inventory tracker spreadsheet"
    wb.properties.description = "Inventory Tracker Studio for replenishment decisions, stock, purchases, sales, reorder alerts, pricing, and snapshots with seeded human-scale capacity."
    wb.properties.keywords = "inventory tracker, stock tracker, replenishment, reorder date, order quantity, purchases, sales, pricing spreadsheet"
    wb.properties.category = "business"


def build_workbook():
    BASE.mkdir(parents=True, exist_ok=True)
    for path in [PRODUCT.parent, IMPORT_COPY.parent, PDF_CHECK.parent, PREVIEW_DIR, RENDER_DIR, DELIVERY_DIR]:
        path.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    start = wb.create_sheet("Start Here")
    dashboard = wb.create_sheet("Dashboard")
    setup = wb.create_sheet("Setup")
    inventory = wb.create_sheet("Inventory")
    purchases_ws = wb.create_sheet("Purchases")
    sales_ws = wb.create_sheet("Sales")
    price_calc = wb.create_sheet("Price Calculator")
    price_list = wb.create_sheet("Price List")
    help_ws = wb.create_sheet("Help")
    lists_ws = wb.create_sheet("Lists")
    build_lists(lists_ws)
    build_start_here(start)
    build_dashboard(dashboard)
    build_setup(setup)
    build_inventory(inventory)
    build_purchases(purchases_ws)
    build_sales(sales_ws)
    build_price_calculator(price_calc)
    build_price_list(price_list)
    build_help(help_ws)
    configure_validation(wb)
    add_comments(wb)
    protect_surfaces(wb)
    workbook_metadata(wb)
    wb.active = wb.sheetnames.index("Start Here")
    wb.save(PRODUCT)
    wb.save(IMPORT_COPY)


def render_pdf_and_previews():
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        env = os.environ.copy()
        env["HOME"] = str(tmp_path)
        subprocess.run(
            [
                "soffice",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(PDF_CHECK.parent),
                str(PRODUCT),
            ],
            check=True,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
    if not PDF_CHECK.exists():
        raise FileNotFoundError(PDF_CHECK)
    doc = fitz.open(str(PDF_CHECK))
    # Pages should map to sheets by order because every sheet is fit-to-page.
    page_names = ["Start Here", "Dashboard", "Setup", "Inventory", "Purchases", "Sales", "Price Calculator", "Price List", "Help"]
    for i in range(min(len(page_names), len(doc))):
        page = doc[i]
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        out = PREVIEW_DIR / f"{i+1:02d}-{page_names[i].lower().replace(' ', '-')}.png"
        pix.save(str(out))


def write_delivery_assets():
    (DELIVERY_DIR / "README-FIRST.md").write_text(
        "Open Inventory-Tracker-Studio.xlsx, then start on Start Here. Type only in yellow cells. Leave formulas, charts, and reorder outputs alone. Inventory shows reorder action, reorder-by date, projected stockout date, days of cover, suggested order quantity, velocity, lead time, and safety stock. This workbook ships with 100 inventory rows, 300 purchase rows, 500 sales rows, 100 calculator rows, and 100 price-list rows.\n",
        encoding="utf-8",
    )
    (DELIVERY_DIR / "Quick-Start-and-FAQ.md").write_text(
        "\n".join(
            [
                "Quick Start",
                "",
                "1. Open Setup and confirm review date, lookback days, lead time, safety stock days, pack size, fees, and target margin.",
                "2. Fill Inventory opening stock, reorder points, and unit costs.",
                "3. Add purchase and sales rows as they happen.",
                "4. Read Inventory reorder action, reorder-by date, projected stockout date, days cover, and suggested order quantity.",
                "5. Use Price Calculator before listing a product and save successful prices in Price List.",
                "",
                "FAQ",
                "- This workbook does not connect to banks or live inventory services.",
                "- Imported protection behavior can vary by spreadsheet editor.",
                "- Keep one clean copy before making heavy edits.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def main():
    build_workbook()
    render_pdf_and_previews()
    write_delivery_assets()
    print(PRODUCT)


if __name__ == "__main__":
    main()
